import io, csv
from datetime import datetime
from django.http import HttpResponse
from django.contrib.auth.models import User
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from accounts.models import Profile

# ── Ranglar ──────────────────────────────────────────────────────────────
GOLD='C9A84C'; DARK='0A0A0F'; HEADER='1E1E2A'; BG='16161F'; ALT='111118'
TEXT='E8E8F0'; MUTED='8888AA'; GREEN='52C078'; RED='E05252'

def _f(color): return PatternFill("solid", fgColor=color)
def _border():
    s=Side(style='thin',color="2A2A3A")
    return Border(left=s,right=s,top=s,bottom=s)
def _font(bold=False,size=10,color=TEXT): return Font(name="Arial",bold=bold,size=size,color=color)
def _align(h="center",v="center",indent=0,wrap=False):
    return Alignment(horizontal=h,vertical=v,indent=indent,wrap_text=wrap)

def _daraja(rating):
    if rating>=2400: return "♛ Grandmaster"
    if rating>=2200: return "♝ Master"
    if rating>=2000: return "♞ Expert"
    if rating>=1800: return "♜ Advanced"
    return "♟ Intermediate"


def _get_players(limit="all", status_filter="all", rating_filter="all"):
    qs = Profile.objects.select_related('user').order_by('-rating')
    if status_filter == "active":   qs = qs.filter(user__is_active=True)
    elif status_filter == "inactive": qs = qs.filter(user__is_active=False)
    if rating_filter != "all": qs = qs.filter(rating__gte=int(rating_filter))
    if limit != "all": qs = qs[:int(limit)]
    return list(qs)


# ── Sheet 1: Reyting jadvali ─────────────────────────────────────────────
def _sheet_main(ws, players):
    ws.merge_cells("A1:L1")
    ws["A1"] = f"♛  ChessMaster UZ — Reyting Jadvali  |  {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A1"].font = Font(name="Arial",bold=True,size=16,color=GOLD)
    ws["A1"].fill = _f(DARK); ws["A1"].alignment = _align(); ws.row_dimensions[1].height=42

    heads = ["#","Foydalanuvchi","To'liq ism","Reyting","G'alaba","Mag'lubiyat",
             "Durang","Jami o'yin","G'alaba %","Daraja","Holat","A'zo bo'lgan"]
    widths=[5,22,24,11,12,14,10,13,11,20,13,16]
    for ci,(h,w) in enumerate(zip(heads,widths),1):
        c=ws.cell(row=2,column=ci,value=h)
        c.font=_font(bold=True,size=10,color=GOLD); c.fill=_f(HEADER)
        c.alignment=_align(wrap=True); c.border=_border()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[2].height=28

    for ri,p in enumerate(players,3):
        u=p.user; rf=_f(ALT) if ri%2==0 else _f(BG)
        wp=round(p.wins/p.games_played*100,1) if p.games_played else 0
        row=[ri-2,u.username,u.get_full_name() or "—",p.rating,
             p.wins,p.losses,p.draws,p.games_played,wp/100,
             _daraja(p.rating),"Aktiv" if u.is_active else "Bloklangan",
             u.date_joined.strftime("%d.%m.%Y") if u.date_joined else "—"]
        for ci,val in enumerate(row,1):
            c=ws.cell(row=ri,column=ci,value=val); c.fill=rf; c.border=_border(); c.alignment=_align()
            c.font=_font(size=10,color=TEXT)
            rank=ri-2
            if ci==1:
                if rank==1: c.font=Font(name="Arial",bold=True,size=11,color="FFD700")
                elif rank==2: c.font=Font(name="Arial",bold=True,size=11,color="C0C0C0")
                elif rank==3: c.font=Font(name="Arial",bold=True,size=11,color="CD7F32")
                else: c.font=_font(size=10,color=MUTED)
            elif ci==4: c.font=_font(bold=True,size=11,color=GOLD)
            elif ci==5: c.font=_font(bold=True,size=10,color=GREEN)
            elif ci==6: c.font=_font(bold=True,size=10,color=RED)
            elif ci==9:
                c.number_format="0.0%"
                c.font=_font(bold=True,size=10,color=GREEN if wp>=60 else(GOLD if wp>=40 else RED))
            elif ci==10: c.font=_font(bold=True,size=10,color=GOLD)
            elif ci==11: c.font=_font(bold=True,size=10,color=GREEN if u.is_active else RED)
            elif ci in(2,3): c.alignment=_align("left",indent=1)
        ws.row_dimensions[ri].height=22

    tr=len(players)+3
    ws.merge_cells(f"A{tr}:C{tr}")
    ws[f"A{tr}"]="JAMI / O'RTACHA"; ws[f"A{tr}"].font=_font(bold=True,size=10,color=GOLD)
    ws[f"A{tr}"].fill=_f(HEADER); ws[f"A{tr}"].alignment=_align(); ws[f"A{tr}"].border=_border()
    for col,formula in [("D",f"=AVERAGE(D3:D{tr-1})"),("E",f"=SUM(E3:E{tr-1})"),
                         ("F",f"=SUM(F3:F{tr-1})"),("G",f"=SUM(G3:G{tr-1})"),
                         ("H",f"=SUM(H3:H{tr-1})"),("I",f"=AVERAGE(I3:I{tr-1})")]:
        c=ws[f"{col}{tr}"]; c.value=formula; c.font=_font(bold=True,size=10,color=GOLD)
        c.fill=_f(HEADER); c.border=_border(); c.alignment=_align()
        if col=="I": c.number_format="0.0%"
    for col in["J","K","L"]:
        c=ws[f"{col}{tr}"]; c.fill=_f(HEADER); c.border=_border()
    ws.row_dimensions[tr].height=26; ws.freeze_panes="A3"


# ── Sheet 2: Top 10 + diagramma ─────────────────────────────────────────
def _sheet_top10(ws, players):
    ws.merge_cells("A1:G1"); ws["A1"]="🏆  Top 10 O'yinchilar"
    ws["A1"].font=Font(name="Arial",bold=True,size=15,color=GOLD)
    ws["A1"].fill=_f(DARK); ws["A1"].alignment=_align(); ws.row_dimensions[1].height=38
    heads=["O'yinchi","Reyting","G'alaba %","G'alaba","Mag'lubiyat","Durang","Daraja"]
    widths=[22,12,13,12,14,10,20]
    for ci,(h,w) in enumerate(zip(heads,widths),1):
        c=ws.cell(row=2,column=ci,value=h); c.font=_font(bold=True,size=10,color=GOLD)
        c.fill=_f(HEADER); c.alignment=_align(); c.border=_border()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[2].height=28
    for ri,p in enumerate(players[:10],3):
        rf=_f(ALT) if ri%2==0 else _f(BG)
        wp=round(p.wins/p.games_played*100,1) if p.games_played else 0
        row=[p.user.username,p.rating,wp/100,p.wins,p.losses,p.draws,_daraja(p.rating)]
        for ci,val in enumerate(row,1):
            c=ws.cell(row=ri,column=ci,value=val); c.fill=rf; c.border=_border(); c.alignment=_align()
            c.font=_font(size=10,color=TEXT)
            if ci==1: c.alignment=_align("left",indent=1); c.font=_font(bold=True,size=10,color=TEXT)
            if ci==2: c.font=_font(bold=True,size=11,color=GOLD)
            if ci==3: c.number_format="0.0%"; c.font=_font(bold=True,size=10,color=GREEN if wp>=50 else RED)
            if ci==4: c.font=_font(bold=True,size=10,color=GREEN)
            if ci==5: c.font=_font(bold=True,size=10,color=RED)
            if ci==7: c.font=_font(bold=True,size=10,color=GOLD)
        ws.row_dimensions[ri].height=22
    if players[:10]:
        chart=BarChart(); chart.type="bar"
        chart.title="Top 10 — Reyting"; chart.y_axis.title="Reyting"; chart.style=10
        chart.width=24; chart.height=14
        data=Reference(ws,min_col=2,min_row=2,max_row=min(12,len(players[:10])+2))
        cats=Reference(ws,min_col=1,min_row=3,max_row=min(12,len(players[:10])+2))
        chart.add_data(data,titles_from_data=True); chart.set_categories(cats)
        if chart.series: chart.series[0].graphicalProperties.solidFill=GOLD
        ws.add_chart(chart,"A14")
    ws.freeze_panes="A3"


# ── Sheet 3: Statistika ──────────────────────────────────────────────────
def _sheet_stats(ws, players):
    ws.merge_cells("A1:B1"); ws["A1"]="📊  Platforma Statistikasi"
    ws["A1"].font=Font(name="Arial",bold=True,size=15,color=GOLD)
    ws["A1"].fill=_f(DARK); ws["A1"].alignment=_align(); ws.row_dimensions[1].height=38
    ws.column_dimensions["A"].width=32; ws.column_dimensions["B"].width=20
    total=len(players); active=sum(1 for p in players if p.user.is_active)
    total_games=sum(p.games_played for p in players)
    avg_r=int(sum(p.rating for p in players)/total) if total else 0
    stats=[("Jami o'yinchilar",total),("Aktiv o'yinchilar",active),
           ("Bloklangan",total-active),("O'rtacha reyting",avg_r),
           ("Eng yuqori reyting",max((p.rating for p in players),default=0)),
           ("Eng past reyting",min((p.rating for p in players),default=0)),
           ("Jami o'yinlar",total_games),
           ("Hisobot sanasi",datetime.now().strftime("%d.%m.%Y %H:%M"))]
    for ci,h in enumerate(["Ko'rsatkich","Qiymat"],1):
        c=ws.cell(row=2,column=ci,value=h); c.font=_font(bold=True,size=10,color=GOLD)
        c.fill=_f(HEADER); c.border=_border(); c.alignment=_align(); ws.row_dimensions[2].height=26
    for ri,(lbl,val) in enumerate(stats,3):
        rf=_f(ALT) if ri%2==0 else _f(BG)
        for ci,v in enumerate([lbl,val],1):
            c=ws.cell(row=ri,column=ci,value=v); c.fill=rf; c.border=_border()
            c.alignment=_align("left" if ci==1 else "center",indent=1 if ci==1 else 0)
            c.font=_font(bold=(ci==2),size=10 if ci==1 else 11,color=TEXT if ci==1 else GOLD)
        ws.row_dimensions[ri].height=24
    ws.cell(row=12,column=1,value="DARAJA TAQSIMOTI").font=_font(bold=True,size=11,color=GOLD)
    ws.cell(row=12,column=1).fill=_f(HEADER); ws.cell(row=12,column=1).border=_border()
    ws.cell(row=12,column=2,value="Soni").font=_font(bold=True,size=11,color=GOLD)
    ws.cell(row=12,column=2).fill=_f(HEADER); ws.cell(row=12,column=2).border=_border()
    ws.row_dimensions[12].height=26
    levels=[("♛ Grandmaster (2400+)",lambda p:p.rating>=2400),
            ("♝ Master (2200-2399)", lambda p:2200<=p.rating<2400),
            ("♞ Expert (2000-2199)", lambda p:2000<=p.rating<2200),
            ("♜ Advanced (1800-1999)",lambda p:1800<=p.rating<2000),
            ("♟ Intermediate (<1800)",lambda p:p.rating<1800)]
    for ri,(lbl,fn) in enumerate(levels,13):
        cnt=sum(1 for p in players if fn(p)); rf=_f(ALT) if ri%2==0 else _f(BG)
        for ci,v in enumerate([lbl,cnt],1):
            c=ws.cell(row=ri,column=ci,value=v); c.fill=rf; c.border=_border()
            c.alignment=_align("left" if ci==1 else "center",indent=1 if ci==1 else 0)
            c.font=_font(bold=(ci==2),size=10 if ci==1 else 11,color=TEXT if ci==1 else GOLD)
        ws.row_dimensions[ri].height=22


# ── Sheet 4: O'yinlar tarixi ─────────────────────────────────────────────
def _sheet_games(ws):
    ws.merge_cells("A1:H1"); ws["A1"]="♟  O'yinlar Tarixi (oxirgi 500)"
    ws["A1"].font=Font(name="Arial",bold=True,size=15,color=GOLD)
    ws["A1"].fill=_f(DARK); ws["A1"].alignment=_align(); ws.row_dimensions[1].height=38
    heads=["#","Oq o'yinchi","Qora o'yinchi","Natija","Oq reyting","Qora reyting","Vaqt","Sana"]
    widths=[5,20,20,16,13,13,12,20]
    for ci,(h,w) in enumerate(zip(heads,widths),1):
        c=ws.cell(row=2,column=ci,value=h); c.font=_font(bold=True,size=10,color=GOLD)
        c.fill=_f(HEADER); c.alignment=_align(); c.border=_border()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[2].height=28
    try:
        from chess.models import Game
        from django.db.models import Q as DQ
        games=Game.objects.filter(
            status__in=['white_wins','black_wins','draw']
        ).select_related('white_player','black_player',
                         'white_player__profile','black_player__profile'
        ).order_by('-created_at')[:500]
        rmap={'white_wins':"Oq g'alaba",'black_wins':"Qora g'alaba",'draw':"Durang"}
        for ri,g in enumerate(games,3):
            rf=_f(ALT) if ri%2==0 else _f(BG)
            row=[ri-2,g.white_player.username,g.black_player.username,
                 rmap.get(g.status,g.status),
                 g.white_rating_before,g.black_rating_before,
                 g.time_control,
                 g.created_at.strftime("%d.%m.%Y %H:%M") if g.created_at else "—"]
            for ci,val in enumerate(row,1):
                c=ws.cell(row=ri,column=ci,value=val); c.fill=rf; c.border=_border()
                c.alignment=_align(); c.font=_font(size=10,color=TEXT)
                if ci==4:
                    if "Oq" in str(val): c.font=_font(bold=True,size=10,color="E0E0E0")
                    elif "Qora" in str(val): c.font=_font(bold=True,size=10,color=GOLD)
                    else: c.font=_font(bold=True,size=10,color=MUTED)
            ws.row_dimensions[ri].height=20
    except Exception as e:
        ws.cell(row=3,column=1,value=f"Xato: {e}").font=_font(size=10,color=RED)
    ws.freeze_panes="A3"


# ── Sheet 5: Admin/Superuser maxfiy ─────────────────────────────────────
def _sheet_admin(ws, players):
    ws.merge_cells("A1:I1"); ws["A1"]="🛡  Admin Hisoboti — MAXFIY"
    ws["A1"].font=Font(name="Arial",bold=True,size=15,color="FF4444")
    ws["A1"].fill=PatternFill("solid",fgColor="1A0000"); ws["A1"].alignment=_align(); ws.row_dimensions[1].height=38
    ws.merge_cells("A2:I2")
    ws["A2"]=f"⚠  Maxfiy. Faqat admin/superuser.  |  {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    ws["A2"].font=Font(name="Arial",italic=True,size=9,color="FF8888")
    ws["A2"].fill=PatternFill("solid",fgColor="120000"); ws["A2"].alignment=_align(); ws.row_dimensions[2].height=20
    heads=["#","Username","To'liq ism","Email","Rol","Reyting","Qo'shilgan","So'nggi kirish","Holat"]
    widths=[5,20,22,30,13,10,14,20,13]
    for ci,(h,w) in enumerate(zip(heads,widths),1):
        c=ws.cell(row=3,column=ci,value=h)
        c.font=Font(name="Arial",bold=True,size=10,color="FF4444")
        c.fill=PatternFill("solid",fgColor="1E0505"); c.border=_border(); c.alignment=_align()
        ws.column_dimensions[get_column_letter(ci)].width=w
    ws.row_dimensions[3].height=26
    for ri,p in enumerate(players,4):
        u=p.user; rf=PatternFill("solid",fgColor="120000" if ri%2==0 else "0A0000")
        if u.is_superuser: rol="👑 Superuser"
        elif u.is_staff: rol="🛡 Admin"
        else: rol="♟ O'yinchi"
        row=[ri-3,u.username,u.get_full_name() or "—",u.email,rol,p.rating,
             u.date_joined.strftime("%d.%m.%Y") if u.date_joined else "—",
             u.last_login.strftime("%d.%m.%Y %H:%M") if u.last_login else "Hech qachon",
             "Aktiv" if u.is_active else "Bloklangan"]
        for ci,val in enumerate(row,1):
            c=ws.cell(row=ri,column=ci,value=val); c.fill=rf; c.border=_border()
            c.alignment=_align(); c.font=_font(size=10,color=TEXT)
            if ci==4: c.alignment=_align("left",indent=1)
            if ci==5:
                if "Superuser" in str(val): c.font=Font(name="Arial",bold=True,size=10,color="FF4444")
                elif "Admin" in str(val): c.font=_font(bold=True,size=10,color=GOLD)
            if ci==6: c.font=_font(bold=True,size=10,color=GOLD)
            if ci==9: c.font=_font(bold=True,size=10,color=GREEN if u.is_active else RED)
        ws.row_dimensions[ri].height=22
    ws.freeze_panes="A4"


# ── Asosiy funksiyalar ───────────────────────────────────────────────────
def build_excel_response(request):
    sheets  = set(request.POST.get('sheets','main,top10,stats,games,admin').split(','))
    limit   = request.POST.get('limit','all')
    sf      = request.POST.get('status_filter','all')
    rf      = request.POST.get('rating_filter','all')
    if not request.user.is_superuser: sheets.discard('admin')
    players = _get_players(limit, sf, rf)
    wb = Workbook(); first = True
    sheet_map = [
        ('main',   "Reyting Jadvali", _sheet_main),
        ('top10',  "Top 10 Tahlil",   _sheet_top10),
        ('stats',  "Statistika",      _sheet_stats),
        ('games',  "O'yinlar Tarixi", _sheet_games),
        ('admin',  "Admin Hisoboti",  _sheet_admin),
    ]
    for key, title, fn in sheet_map:
        if key not in sheets: continue
        if first:
            ws = wb.active; ws.title = title; first = False
        else:
            ws = wb.create_sheet(title)
        if key in ('main','top10','stats','admin'): fn(ws, players)
        else: fn(ws)
    if first:
        ws = wb.active; ws.title = "Reyting Jadvali"; _sheet_main(ws, _get_players())
    fname = f"ChessMaster_Reyting_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    resp = HttpResponse(buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    return resp


def build_csv_response(request):
    fname = f"ChessMaster_Reyting_{datetime.now().strftime('%Y%m%d')}.csv"
    resp = HttpResponse(content_type='text/csv; charset=utf-8-sig')
    resp['Content-Disposition'] = f'attachment; filename="{fname}"'
    w = csv.writer(resp)
    w.writerow(['#','Username',"To'liq ism",'Email','Reyting',
                "G'alaba","Mag'lubiyat","Durang","Jami o'yin","G'alaba %",'Rol','Holat'])
    for i,p in enumerate(_get_players(),1):
        u=p.user; wp=f"{round(p.wins/p.games_played*100,1)}%" if p.games_played else "0%"
        if u.is_superuser: rol="Superuser"
        elif u.is_staff: rol="Admin"
        else: rol="O'yinchi"
        w.writerow([i,u.username,u.get_full_name() or "",
                    u.email if request.user.is_superuser else "***",
                    p.rating,p.wins,p.losses,p.draws,p.games_played,wp,rol,
                    "Aktiv" if u.is_active else "Bloklangan"])
    return resp
